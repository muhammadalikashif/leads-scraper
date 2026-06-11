"""
Google Maps Business Data Scraper (generic, max results)
=========================================================
Scrapes for every search result returned by Google Maps:
  - Business details  : name, branch, category, rating, review count, address, phone, website, overview
  - About tab         : clean structured attributes (Accessibility, Payments, etc.)

OUTPUT — 1 CSV file per run:
  gmaps_businesses_TIMESTAMP.csv

The scraper is country-agnostic:
  * Add queries to QUERIES (one per search; Arabic + English both supported)
  * Optionally set PHONE_REGION = "PK" / "US" / "SA" for friendly phone grouping

To collect the maximum number of results per query, set
  INFINITE_SCROLL = True
  MAX_RESULTS_PER_QUERY = 0
The feed will be scrolled until no new results load (capped by
MAX_SCROLL_ATTEMPTS).
"""

import csv
import os
import re
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import List, Optional, Set
from urllib.parse import quote_plus, urlparse, parse_qs, unquote, urlunparse
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

# Force UTF-8 on Windows consoles so Arabic / Urdu / etc. print correctly.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, OSError):
        pass

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



# =============================================================================
# CONFIGURATION
# =============================================================================

# === Queries ===
# One entry = one search. Add more strings to scrape additional terms.
# Each entry becomes its own Google Maps search; results are deduplicated
# across queries by Maps URL, name, phone, and address.
QUERIES: List[str] = ["Electronics store Riyadh"]
# === Scroll / result limits ===
# Set MAX_RESULTS_PER_QUERY = 0 to collect all results Google returns for a query.
# When INFINITE_SCROLL is True the feed is scrolled until no new results load
# (capped by MAX_SCROLLS as a safety net). When False, the feed is scrolled
# exactly MAX_SCROLLS times.
MAX_RESULTS_PER_QUERY: int      = 0
INFINITE_SCROLL:       bool     = True
MAX_SCROLLS:           int      = 100        # fixed scrolls when INFINITE_SCROLL=False
SCROLL_STAGNANT_LIMIT: int      = 5         # stop after N no-growth scrolls (infinite mode)
SCROLL_PAUSE:          float    = 2.0

# === Detail page ===
DETAIL_PAUSE:          float    = 3.0
DETAIL_RETRIES:        int      = 3

# === Browser ===
HEADLESS:              bool     = True
PAGE_LOAD_TIMEOUT:     int      = 45
LANG:                  str      = "en"

# === Phone formatting ===
# Set PHONE_REGION to an ISO 2-letter code (e.g. "SA", "PK", "US") to enable
# friendly grouping like "+966 50 123 4567". Set to "" for a generic E.164
# (+CCxxxxxxxxxx) output with no region-specific grouping.
PHONE_REGION:          str      = ""

TIMESTAMP   = datetime.now().strftime("%Y%m%d_%H%M%S")
BIZ_FILE = f"/app/output/gmaps_businesses_{TIMESTAMP}.csv"
BIZ_XLSX = f"/app/output/gmaps_businesses_{TIMESTAMP}.xlsx"


# =============================================================================
# REGEX CLEANERS
# =============================================================================

# Extended emoji pattern: covers emoticons, misc symbols, food, transport,
# flags, dingbats, supplemental symbols, and common Unicode emoji
EMOJI_RE = re.compile(
    "["
    "\U0001F600-\U0001F64F"  # emoticons
    "\U0001F300-\U0001F5FF"  # misc symbols & pictographs
    "\U0001F680-\U0001F6FF"  # transport & map
    "\U0001F1E0-\U0001F1FF"  # flags
    "\U00002702-\U000027B0"  # dingbats
    "\U000024C2-\U0001F251"  # enclosed chars
    "\U0001F900-\U0001F9FF"  # supplemental symbols
    "\U0001FA00-\U0001FA6F"  # chess symbols
    "\U0001FA70-\U0001FAFF"  # symbols extended-A
    "\U00002600-\U000026FF"  # misc symbols (☀★ etc)
    "\U0000FE00-\U0000FE0F"  # variation selectors
    "\U0000200D"             # ZWJ
    "\U00002B50"             # star
    "\U000023F0-\U000023FA"  # misc technical
    "\U0000FE0F"             # VS16
    "\ue000-\uf8ff"          # Google private-use icon chars
    "]+",
    flags=re.UNICODE,
)

# Plus Code pattern: 2-4 chars + "+" + 2-4 chars, anchored to word boundaries
# (so we don't eat fragments of normal text like "C++" or postal codes).
# An optional trailing Urdu/Arabic comma or whitespace is consumed.
PLUS_CODE_RE = re.compile(
    r"(?<![A-Za-z0-9])[A-Z0-9]{2,4}\+[A-Z0-9]{2,4}[،,\s]*(?![A-Za-z0-9])",
    re.IGNORECASE,
)


def serve_debug_output_if_enabled() -> None:
    """
    Temporary helper to view /app/output files from the browser.

    Enable with:
      SERVE_DEBUG_FILES=1

    Easy to remove later:
      1. Delete this method.
      2. Delete the call after main().
      3. Remove SERVE_DEBUG_FILES from Coolify.
    """
    if os.getenv("SERVE_DEBUG_FILES", "0") != "1":
        return

    output_dir = os.path.dirname(BIZ_FILE) or "/app/output"
    os.makedirs(output_dir, exist_ok=True)
    os.chdir(output_dir)

    port = int(os.getenv("PORT", "3000"))

    print("=" * 70)
    print(f"Serving debug output directory: {output_dir}")
    print(f"Listening on: 0.0.0.0:{port}")
    print("=" * 70)

    server = ThreadingHTTPServer(("0.0.0.0", port), SimpleHTTPRequestHandler)
    server.serve_forever()

def clean_text(value: Optional[str]) -> str:
    """Strip whitespace, newlines, tabs, icon chars, and emojis."""
    if not value:
        return ""
    value = EMOJI_RE.sub("", value)
    value = re.sub(r"[\r\n\t]+", " ", value)             # newlines/tabs -> space
    value = re.sub(r"\s{2,}", " ", value)                # multiple spaces -> one
    return value.strip()


def clean_phone(value: str) -> str:
    """Normalize phone to E.164-style: +923245258655 or keep formatted: +92 324 5258655."""
    if not value:
        return ""
    # Strip everything except digits and leading +
    value = re.sub(r"(?i)^phone:\s*", "", value)
    digits = re.sub(r"[^\d+]", "", value)
    # Ensure leading +
    if digits and not digits.startswith("+"):
        digits = "+" + digits
    return digits


def clean_phone_display(value: str, region: str = "") -> str:
    """
    Format phone for display.
      region="PK"  -> "+92 324 5258655" / "+92 51 111 532 532"  (Pakistan rules)
      region="US"  -> "+1 234 555 0123"
      region=""    -> generic E.164: "+CCxxxxxxxxxx" with + and digits only
    """
    raw = clean_phone(value)
    if not raw:
        return ""
    digits = raw.lstrip("+")

    if region.upper() == "PK":
        # Strip a leading trunk "0" so local format becomes international.
        if digits.startswith("0"):
            digits = "92" + digits[1:]
        m = re.match(r"^92(\d{3})(\d{7})$", digits)
        if m:
            return f"+92 {m.group(1)} {m.group(2)}"
        m = re.match(r"^92(\d{2})(\d{7,9})$", digits)
        if m:
            return f"+92 {m.group(1)} {m.group(2)}"
        return "+" + digits

    if region.upper() == "US":
        # Strip leading "1" trunk prefix if present.
        if digits.startswith("1") and len(digits) == 11:
            digits = digits[1:]
        m = re.match(r"^(\d{3})(\d{3})(\d{4})$", digits)
        if m:
            return f"+1 {m.group(1)} {m.group(2)} {m.group(3)}"
        return "+1" + digits if not digits.startswith("1") else "+" + digits

    if region.upper() == "SA":
        # Saudi Arabia: drop leading trunk "0", ensure country code 966
        if digits.startswith("0"):
            digits = "966" + digits[1:]
        # Mobile: 966 5X XXX XXXX
        m = re.match(r"^966(5\d)(\d{3})(\d{4})$", digits)
        if m:
            return f"+966 {m.group(1)} {m.group(2)} {m.group(3)}"
        # Landline: 966 X XXX XXXX (Makkah area = 2, Riyadh = 1, etc.)
        m = re.match(r"^966(\d{1,2})(\d{3})(\d{4})$", digits)
        if m:
            return f"+966 {m.group(1)} {m.group(2)} {m.group(3)}"
        return "+" + digits

    # Generic: keep + prefix, digits only
    return "+" + digits


def clean_rating(value: str) -> str:
    """Extract numeric rating e.g. '4.3'."""
    if not value:
        return ""
    m = re.search(r"\d+(?:\.\d+)?", value)
    return m.group(0) if m else ""


def rating_float(value: str) -> str:
    """Return rating as a float string for sorting, e.g. '4.3'."""
    r = clean_rating(value)
    if not r:
        return ""
    try:
        return str(float(r))
    except ValueError:
        return ""


def clean_review_count(value: str) -> str:
    """Extract plain number from '(2,228 reviews)' -> '2228'."""
    if not value:
        return ""
    digits = re.sub(r"[^\d]", "", value)
    return digits if digits else ""





def clean_maps_url(url: str) -> str:
    """
    Clean Maps URL: keep the place name path but strip:
    - Coordinate params (/@lat,lng,zoom,/data=!3m1!4b1)
    - Query parameters and fragments
    Example: .../place/KFC+F10/@33.7,73.05,17z/data=!3m1 -> .../place/KFC+F10
    """
    if not url:
        return ""
    # Strip everything after /@ (coordinates) or /data= or ? or #
    url = re.sub(r"/@.*$", "", url)
    url = re.sub(r"/data=.*$", "", url)
    parsed = urlparse(url)
    clean = urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", "", ""))
    return clean


def extract_branch_name(name: str, address: str) -> str:
    """
    Extract branch/location identifier from business name and address.
    Generic: handles single-letter sectors (E11, F-10), multi-letter (G-15),
    'Markaz' (F-11 Markaz), 'Block' (Block B), 'Phase' (Phase 6), etc.
    """
    # 1) name-based: "KFC | E11", "KFC F10", "KFC - D12", "KFC, G-15 Markaz"
    for pat in [
        r"[\|\-,]\s*([A-Z]{1,2}[\s\-]?\d{1,3}[A-Z]?)\b",                 # KFC | E11 / G-15
        r"\b([A-Z]{1,2}\d{1,3}[A-Z]?)\b",                              # bare E11, G15B
    ]:
        m = re.search(pat, name)
        if m:
            token = re.sub(r"[\s\-]", "", m.group(1)).upper()
            # Skip very short tokens that look like noise
            if len(token) >= 2:
                return token

    # 2) address-based: "F-11 Markaz", "Block B", "Phase 6", "Sector 18"
    for pat in [
        r"\b([A-Z]{1,2})\s*[\-]?\s*(\d{1,3})\b",                       # F-11, E 18
        r"\bBlock\s+([A-Z])\b",
        r"\bPhase\s+(\d+)\b",
        r"\bSector\s+(\d+)\b",
    ]:
        m = re.search(pat, address, re.IGNORECASE)
        if m:
            if len(m.groups()) == 2:
                return f"{m.group(1).upper()}{m.group(2)}"
            return m.group(1).upper()

    # 3) Markaz / area fallback
    m = re.search(r"\b([A-Z]{1,2}\d{1,3}[A-Z]?)\s+Markaz\b", address, re.IGNORECASE)
    if m:
        return re.sub(r"[\s\-]", "", m.group(1)).upper()
    m = re.search(r"Markaz\s+([A-Z]{1,2}\d{1,3}[A-Z]?)", address, re.IGNORECASE)
    if m:
        return re.sub(r"[\s\-]", "", m.group(1)).upper()

    return ""


INVALID_DOMAINS = ["google.", "maps.google.", "g.page", "goo.gl", "gstatic.", "googleusercontent."]

def clean_website(url: str) -> str:
    """
    Clean and normalize website URL:
    - Unwrap Google redirect URLs
    - Normalize to https
    - Strip www. prefix
    - Strip trailing slashes and fragments
    - Reject Google/internal domains
    """
    if not url:
        return ""
    # Unwrap Google redirect
    parsed = urlparse(url)
    if "google" in parsed.netloc and parsed.path == "/url":
        params = parse_qs(parsed.query)
        if "q" in params:
            url = unquote(params["q"][0])
    parsed = urlparse(url)
    domain = parsed.netloc.lower().replace("www.", "")
    if not domain:
        return ""
    if any(bad in domain for bad in INVALID_DOMAINS):
        return ""
    # Normalize: always https, no www, no trailing slash, no fragment
    clean_path = parsed.path.rstrip("/")
    return urlunparse(("https", domain, clean_path, "", "", ""))


def clean_address(address: str) -> str:
    """
    Clean address:
    - Remove Plus Codes (e.g. 'M2V7+P6V')
    - Remove Urdu/Arabic comma (،)
    - Normalize whitespace
    """
    if not address:
        return ""
    # Remove Plus Codes
    address = PLUS_CODE_RE.sub("", address)
    # Remove Urdu comma
    address = re.sub(r"،", ",", address)
    # Collapse whitespace
    address = re.sub(r"\s{2,}", " ", address)
    # Clean leading/trailing punctuation artifacts
    address = re.sub(r"^[,\s]+|[,\s]+$", "", address)
    return address.strip()




def clean_about(raw: str) -> str:
    """
    Transform messy about text into:
    Accessibility: item1, item2 | Service options: item1, item2 | ...
    """
    if not raw:
        return ""
    # Remove icon characters and emojis
    raw = EMOJI_RE.sub("", raw)
    # Split on || to remove duplicated blocks
    parts = [p.strip() for p in raw.split("||") if p.strip()]
    seen_keys = set()
    unique = []
    for p in parts:
        key = re.sub(r"\s+", " ", p).lower().strip()
        if key not in seen_keys:
            seen_keys.add(key)
            unique.append(p)
    rejoined = " | ".join(unique)
    # Collapse multiple pipes
    rejoined = re.sub(r"(\s*\|\s*){2,}", " | ", rejoined)
    rejoined = re.sub(r"^\s*\|\s*|\s*\|\s*$", "", rejoined)
    rejoined = re.sub(r"\s{2,}", " ", rejoined)
    tokens = [t.strip() for t in rejoined.split("|") if t.strip()]

    SECTION_PAT = re.compile(
        r"^(Accessibility|Service options|Popular for|Offerings|Dining options"
        r"|Amenities|Atmosphere|Crowd|Payments|Children|Parking|Highlights"
        r"|Planning|Health & safety|From the business)$",
        re.IGNORECASE,
    )
    sections: dict = {}
    order: list = []
    current = None
    for token in tokens:
        if SECTION_PAT.match(token):
            current = token.strip()
            if current not in sections:
                sections[current] = []
                order.append(current)
        elif token:
            if current is None:
                current = "General"
                if current not in sections:
                    sections[current] = []
                    order.append(current)
            if token not in sections[current]:
                sections[current].append(token)

    lines = []
    for sec in order:
        items = [i for i in sections[sec] if i]
        if items:
            lines.append(f"{sec}: {', '.join(items)}")
    return " | ".join(lines)


# =============================================================================
# DATA MODELS
# =============================================================================

@dataclass
class Business:
    name:               str = ""
    branch_name:        str = ""
    category:           str = ""
    rating_numeric:     str = ""
    total_reviews:      str = ""
    total_reviews_int:  str = ""
    address:            str = ""
    phone:              str = ""
    website:            str = ""
    maps_url:           str = ""
    overview:           str = ""
    about:              str = ""
    city:               str = ""
    source:             str = "google_maps"


@dataclass
class SearchResult:
    name:     str
    maps_url: str


# =============================================================================
# DRIVER
# =============================================================================

def create_driver():
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-notifications")
    opts.add_argument(f"--lang={LANG}")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-gpu")
    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    return driver


def is_alive(driver) -> bool:
    """A driver is 'alive' only if the CDP session actually responds."""
    try:
        _ = driver.current_url
        ready = driver.execute_script("return document.readyState")
        return bool(ready)
    except Exception:
        return False


def recover(driver):
    print("  Recovering browser...")
    try:
        driver.quit()
    except Exception:
        pass
    time.sleep(2)
    d = create_driver()
    print("  Browser ready.")
    return d


def save_debug_artifacts(driver, label: str = "debug") -> None:
    """
    Temporary debug helper.

    Saves:
      - screenshot PNG
      - page HTML
      - current URL TXT

    Easy to remove later:
      1. Delete this method.
      2. Delete calls to save_debug_artifacts(...).
    """
    try:
        output_dir = os.path.dirname(BIZ_FILE) or "/app/output"
        os.makedirs(output_dir, exist_ok=True)

        safe_label = re.sub(r"[^a-zA-Z0-9_-]+", "_", label).strip("_")[:80]
        debug_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_path = os.path.join(output_dir, f"debug_{debug_ts}_{safe_label}")

        current_url = ""
        try:
            current_url = driver.current_url
        except Exception:
            pass

        try:
            driver.save_screenshot(f"{base_path}.png")
        except Exception as e:
            print(f"  Debug screenshot save failed: {e}")

        try:
            with open(f"{base_path}.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
        except Exception as e:
            print(f"  Debug HTML save failed: {e}")

        try:
            with open(f"{base_path}.txt", "w", encoding="utf-8") as f:
                f.write(f"URL: {current_url}\n")
        except Exception as e:
            print(f"  Debug URL save failed: {e}")

        print(f"  Debug saved: {base_path}.png / .html / .txt")

    except Exception as e:
        print(f"  Debug artifact helper failed: {e}")


# =============================================================================
# WAITING + SCROLLING
# =============================================================================

def wait_results(driver, timeout=25):
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/maps/place/']"))
    )


def wait_detail(driver, timeout=25):
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "h1, button[data-item-id], div[role='main']")
        )
    )


def scroll(driver, selector: str, times: int, pause: float = 1.5) -> int:
    """
    Scroll the results panel. Returns the number of scrolls actually performed.
    Falls back to window scroll only if the panel cannot be located.
    """
    done = 0
    for _ in range(times):
        try:
            panel = driver.find_element(By.CSS_SELECTOR, selector)
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", panel)
            done += 1
        except Exception:
            try:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                done += 1
            except Exception:
                # Browser is unresponsive; let the caller recover.
                break
        time.sleep(pause)
    return done


# =============================================================================
# EXTRACTION HELPERS
# =============================================================================

def first_text(soup, selectors):
    for sel in selectors:
        el = soup.select_one(sel)
        if el:
            t = clean_text(el.get_text(" ", strip=True))
            if t:
                return t
    return ""


def first_attr(soup, selectors, attr):
    for sel in selectors:
        el = soup.select_one(sel)
        if el and el.get(attr):
            v = clean_text(el.get(attr))
            if v:
                return v
    return ""


# =============================================================================
# BUSINESS EXTRACTION
# =============================================================================

def extract_search_results(html: str) -> List[SearchResult]:
    soup = BeautifulSoup(html, "html.parser")
    results, seen = [], set()
    for link in soup.select("a.hfpxzc[href*='/maps/place/'], a[href*='/maps/place/']"):
        url  = clean_text(link.get("href", ""))
        name = clean_text(link.get("aria-label", "")) or clean_text(link.get_text(" ", strip=True))
        if not url or url in seen:
            continue
        seen.add(url)
        results.append(SearchResult(name=name, maps_url=url))
    return results


def extract_about_tab(driver, return_url: str) -> str:
    """
    Click the About tab, scrape structured attributes, then navigate back to
    the place's main detail page (return_url) regardless of success/failure.
    """
    about_text = ""
    try:
        tabs = driver.find_elements(By.CSS_SELECTOR, "button[role='tab']")
        about_tab = next((t for t in tabs if "about" in t.text.lower()), None)
        if not about_tab:
            for t in driver.find_elements(By.CSS_SELECTOR, "button[aria-label*='bout']"):
                about_tab = t
                break
        if about_tab:
            driver.execute_script("arguments[0].click();", about_tab)
            time.sleep(1.5)
            soup = BeautifulSoup(driver.page_source, "html.parser")
            lines = []
            for sec in soup.select("div.iP2t7d, div.OqCZI, div[aria-label] div.fontBodyMedium"):
                t = clean_text(sec.get_text(" | ", strip=True))
                if t and len(t) > 2:
                    lines.append(t)
            hours_el = soup.select_one("div.t39EBf, table.eK4R0e")
            if hours_el:
                h = clean_text(hours_el.get_text(" ", strip=True))
                if h:
                    lines.append("Hours: " + h)
            about_text = clean_about(" || ".join(lines)) if lines else clean_about(
                clean_text(driver.page_source)
            )
    except Exception:
        pass
    finally:
        # Always return to the main detail page so the caller can scrape
        # further fields (or the next iteration can re-load).
        try:
            if return_url and is_alive(driver):
                driver.get(return_url)
        except Exception:
            pass
    return about_text


def extract_business(driver, result: SearchResult, city: str) -> Business:
    soup = BeautifulSoup(driver.page_source, "html.parser")

    name = first_text(soup, ["h1.DUwDvf", "h1", "div[role='main'] h1"]) or result.name
    category = first_text(soup, ["button[jsaction*='pane.rating.category']", "button.DkEaL", "div.DkEaL"])

    raw_rating = first_text(soup, ["span.MW4etd", "div.F7nice span[aria-hidden='true']"])
    rating_numeric = rating_float(raw_rating)

    raw_reviews = first_text(soup, ["span.UY7F9", "button[jsaction*='reviewChart'] span", "span[aria-label*='review']"])
    total_reviews = clean_review_count(raw_reviews)

    addr_raw = first_attr(soup, ["button[data-item-id='address']", "button[data-item-id*='address']"], "aria-label")
    address_raw = re.sub(r"(?i)^address:\s*", "", clean_text(addr_raw)) or \
                  first_text(soup, ["button[data-item-id='address'] div.Io6YTe"])
    address = clean_address(address_raw)

    phone_raw = first_attr(soup, ["button[data-item-id^='phone:tel:']", "button[data-item-id*='phone']"], "aria-label")
    phone = clean_phone_display(phone_raw, region=PHONE_REGION)

    website = ""
    for sel in ["a[data-item-id='authority']", "a[data-item-id*='authority']", "a[aria-label*='Website']"]:
        for link in soup.select(sel):
            w = clean_website(clean_text(link.get("href", "")))
            if w:
                website = w
                break
        if website:
            break

    overview = ""
    for sel in ["div.PYvSYb", "div.xt2b0d", "div[jsname='MZnM8e']", "div.WeS02d span"]:
        el = soup.select_one(sel)
        if el:
            t = clean_text(el.get_text(" ", strip=True))
            if t and len(t) > 10:
                overview = t
                break

    # About tab (navigates back to overview internally)
    about = extract_about_tab(driver, return_url=result.maps_url)

    # Extract branch name
    branch = extract_branch_name(name, address)

    return Business(
        name              = clean_text(name),
        branch_name       = branch,
        category          = clean_text(category),
        rating_numeric    = rating_numeric,
        total_reviews     = total_reviews,
        total_reviews_int = total_reviews,
        address           = address,
        phone             = phone,
        website           = website,
        maps_url          = clean_maps_url(driver.current_url),
        overview          = overview,
        about             = about,
        city              = city,
    )


# =============================================================================
# CSV WRITERS
# =============================================================================

BIZ_FIELDS = [
    "name", "branch_name", "category",
    "rating_numeric", "total_reviews", "total_reviews_int",
    "address", "phone", "website", "maps_url",
    "overview", "about", "city", "source",
]


class CSVWriter:
    def __init__(self, filepath: str, fieldnames: list):
        self.filepath  = filepath
        self.count     = 0
        self._file     = open(filepath, "w", newline="", encoding="utf-8-sig")
        self._writer   = csv.DictWriter(self._file, fieldnames=fieldnames)
        self._writer.writeheader()
        self._file.flush()
        print(f"  {filepath}")

    def write(self, obj):
        """
        Write a single record and durably commit it to disk before returning,
        so a crash mid-run never loses records that were already scraped.
        """
        try:
            self._writer.writerow(asdict(obj))
            self._file.flush()
            os.fsync(self._file.fileno())
            self.count += 1
        except Exception as e:
            print(f"  Write error: {e}")

    def close(self):
        if not self._file.closed:
            self._file.close()
        print(f"  {self.count} rows -> {self.filepath}")


class XLSXWriter:
    """
    Excel writer that stores every string field as an explicit text cell so
    values like '+923071330303' are not silently converted into scientific
    notation (9.23071E+11) on open. Rows are flushed and fsync'd after every
    write for crash safety.
    """

    HEADER_FONT = Font(bold=True)

    def __init__(self, filepath: str, fieldnames: list):
        self.filepath = filepath
        self.count    = 0
        self._fieldnames = fieldnames
        self._wb      = Workbook()
        self._ws      = self._wb.active
        self._ws.title = "Businesses"
        # Bold header
        for col_idx, name in enumerate(fieldnames, 1):
            c = self._ws.cell(row=1, column=col_idx, value=name)
            c.font = self.HEADER_FONT
        self._wb.save(filepath)
        self._file = open(filepath, "r+b")
        self._file.flush()
        os.fsync(self._file.fileno())
        print(f"  {filepath}")

    def write(self, obj):
        try:
            row = [str(v) if v is not None else "" for v in asdict(obj).values()]
            self._ws.append(row)
            # Mark every cell in the new row as text so Excel preserves
            # the exact characters (no scientific notation, no auto-URL,
            # no date coercion). The phone column is the main offender.
            for cell in self._ws[self._ws.max_row]:
                cell.data_type = "s"
                cell.number_format = "@"
            self.count += 1
            self._wb.save(self.filepath)
            self._file.flush()
            os.fsync(self._file.fileno())
        except Exception as e:
            print(f"  XLSX write error: {e}")

    def close(self):
        try:
            self._file.close()
        except Exception:
            pass
        print(f"  {self.count} rows -> {self.filepath}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    driver      = create_driver()
    biz_writer  = CSVWriter(BIZ_FILE, BIZ_FIELDS)
    xlsx_writer = XLSXWriter(BIZ_XLSX, BIZ_FIELDS)

    seen_urls:  Set[str] = set()
    seen_dedup: Set[str] = set()

    all_queries = [(q, q) for q in QUERIES]

    print("=" * 70)
    print("  Google Maps Business Scraper")
    print("=" * 70)
    print(f"  Queries : {len(all_queries)}")
    print(f"  Output  : {BIZ_FILE}")
    print("=" * 70)

    try:
        for q_idx, (city, query) in enumerate(all_queries, 1):
            print(f"\n[{q_idx}/{len(all_queries)}] {query}")

            for attempt in range(1, DETAIL_RETRIES + 1):
                try:
                    driver.get(f"https://www.google.com/maps/search/{quote_plus(query)}")
                    wait_results(driver)

                    # Collect URLs from the feed, scrolling either a fixed number
                    # of times (INFINITE_SCROLL=False) or until growth stops.
                    seen_in_query: Set[str] = set()
                    stagnant = 0
                    max_iters = 10**9 if INFINITE_SCROLL else MAX_SCROLLS
                    for scroll_i in range(max_iters):
                        scroll(driver, "div[role='feed']", 1, pause=SCROLL_PAUSE)
                        urls_now = {
                            r.maps_url
                            for r in extract_search_results(driver.page_source)
                        }
                        new = urls_now - seen_in_query
                        if not new:
                            stagnant += 1
                            if not INFINITE_SCROLL or stagnant >= SCROLL_STAGNANT_LIMIT:
                                break
                        else:
                            stagnant = 0
                            seen_in_query.update(new)
                        if not INFINITE_SCROLL and scroll_i + 1 >= MAX_SCROLLS:
                            break

                    raw = [r for r in extract_search_results(driver.page_source)
                           if r.maps_url in seen_in_query]
                    if MAX_RESULTS_PER_QUERY > 0:
                        raw = raw[:MAX_RESULTS_PER_QUERY]
                    print(f"  Found {len(raw)} results")

                    for idx, result in enumerate(raw, 1):
                        if result.maps_url in seen_urls:
                            continue
                        seen_urls.add(result.maps_url)
                        print(f"  [{idx}] {result.name or '(no name)'}", end=" ", flush=True)

                        success = False
                        for d_att in range(1, DETAIL_RETRIES + 1):
                            try:
                                driver.get(result.maps_url)
                                wait_detail(driver)
                                time.sleep(DETAIL_PAUSE)

                                soup_check = BeautifulSoup(driver.page_source, "html.parser")
                                phone_raw  = first_attr(soup_check, ["button[data-item-id^='phone:tel:']", "button[data-item-id*='phone']"], "aria-label")
                                addr_check = first_attr(soup_check, ["button[data-item-id='address']", "button[data-item-id*='address']"], "aria-label")
                                name_check = first_text(soup_check, ["h1.DUwDvf", "h1"]) or result.name

                                phone_key = re.sub(r"\D", "", phone_raw)
                                addr_key  = re.sub(r"\s+", "", (addr_check or "").lower())[:40]
                                name_key  = re.sub(r"\s+", "", (name_check or "").lower())
                                # Prefer phone; fall back to a name+address fingerprint
                                # so phone-less places are not collapsed together.
                                dedup = f"{name_key}|{phone_key}" if phone_key else f"{name_key}|{addr_key}"
                                if dedup in seen_dedup:
                                    print("duplicate — skipped")
                                    success = True
                                    break
                                seen_dedup.add(dedup)

                                biz = extract_business(driver, result, city)
                                biz_writer.write(biz)
                                xlsx_writer.write(biz)
                                print(f"  rating={biz.rating_numeric}  reviews={biz.total_reviews}  branch={biz.branch_name or '-'}")
                                success = True
                                break

                            except Exception as e:
                                if not is_alive(driver):
                                    print(f"crash (attempt {d_att}/{DETAIL_RETRIES})", end=" ", flush=True)
                                    driver = recover(driver)
                                    # Loop continues; recovered driver is reused
                                else:
                                    print(f"error: {e}")
                                    save_debug_artifacts(driver, f"detail_{q_idx}_{idx}_attempt_{d_att}")
                                    break
                        if not success and is_alive(driver):
                            print("  skipped (unrecoverable)")
                    break

                except Exception as e:
                    # If the exception is a browser/CDP failure, always recover.
                    # Otherwise, retry once with a recovered browser as well.
                    print(f"  Query error (attempt {attempt}/{DETAIL_RETRIES}): {e}")

                    if is_alive(driver):
                        save_debug_artifacts(driver, f"query_{q_idx}_attempt_{attempt}")

                    if attempt < DETAIL_RETRIES or not is_alive(driver):
                        driver = recover(driver)
                        # Loop continues; recovered driver is reused on next attempt
                    else:
                        break

    finally:
        biz_writer.close()
        xlsx_writer.close()
        try:
            if is_alive(driver):
                driver.quit()
        except Exception:
            pass
        print(f"\n{'='*70}")
        print(f"  Businesses : {biz_writer.count}  -> {BIZ_FILE}, {BIZ_XLSX}")
        print(f"{'='*70}")

if __name__ == "__main__":
    main()
    serve_debug_output_if_enabled()